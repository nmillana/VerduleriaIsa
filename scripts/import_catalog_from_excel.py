import argparse
from pathlib import Path

from openpyxl import load_workbook

from verduleria.catalog_meta import normalize_name
from verduleria.env import load_env_file
from verduleria.storage import create_database


SECTION_HEADERS = {
    'frutas': 'frutas',
    'verduras': 'verduras',
    'hierbas y complementos': 'hierbas y complementos',
    'legumbres y otros': 'legumbres y otros',
}


def detect_category(label: str, current: str) -> str:
    text = normalize_name(label).lower()
    for key, category in SECTION_HEADERS.items():
        if key in text:
            return category
    return current


def extract_catalog(path: str) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    products = []
    category = 'frutas'
    for row in worksheet.iter_rows(values_only=True):
        name = row[0]
        price = row[2]
        if isinstance(name, str):
            detected = detect_category(name, category)
            if detected != category or normalize_name(name).lower() in SECTION_HEADERS:
                category = detected
                continue
            normalized_name = normalize_name(name)
            if normalized_name.lower() in {'item', 'despacho', 'total'}:
                continue
        if not isinstance(name, str) or not normalize_name(name):
            continue
        try:
            estimated_price = int(price)
        except (TypeError, ValueError):
            continue
        products.append(
            {
                'category': category,
                'name': normalize_name(name),
                'estimated_price': estimated_price,
            }
        )
    return products


def main() -> None:
    parser = argparse.ArgumentParser(description='Importa catálogo desde un Excel semanal')
    parser.add_argument('excel_path', help='Ruta del archivo .xlsx')
    parser.add_argument(
        '--deactivate-missing',
        action='store_true',
        help='Desactiva productos activos que no aparezcan en el Excel nuevo',
    )
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent.parent
    load_env_file(base_dir / '.env')
    catalog = extract_catalog(args.excel_path)
    db = create_database(base_dir)
    db.initialize()
    result = db.sync_catalog(catalog, deactivate_missing=args.deactivate_missing)
    print(
        f"Importación lista. Insertados: {result['inserted']}, "
        f"actualizados: {result['updated']}, desactivados: {result['deactivated']}"
    )


if __name__ == '__main__':
    main()
