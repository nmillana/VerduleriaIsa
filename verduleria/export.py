"""
Exportación de datos a Excel.
Genera archivos Excel con consolidados de pedidos.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_weekly_consolidation_to_excel(consolidation_data: dict) -> bytes:
    """
    Exportar consolidado semanal a archivo Excel.

    Args:
        consolidation_data: Dict con estructura:
            {
                'semana_1': {
                    'tomate': {'cantidad': 5.5, 'precio_unitario': 500, 'total': 2750},
                    ...
                },
                'semana_2': {...}
            }

    Returns:
        Bytes del archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    total_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = '#,##0'
    number_format = '0.00'

    # Encabezados generales
    ws['A1'] = "CONSOLIDADO SEMANAL DE PEDIDOS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    row = 3

    # Iterar por semanas
    for week_label, products in consolidation_data.items():
        # Encabezado de semana
        ws[f'A{row}'] = week_label
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        # Encabezados de tabla
        headers = ['Producto', 'Cantidad', 'Precio Unitario', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row += 1

        # Datos de productos
        subtotal_semana = 0
        for product_name in sorted(products.keys()):
            product_data = products[product_name]

            ws[f'A{row}'] = product_name
            ws[f'B{row}'] = product_data['cantidad']
            ws[f'C{row}'] = product_data['precio_unitario']
            ws[f'D{row}'] = product_data['total']

            # Aplicar estilos y formatos
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws[f'B{row}'].number_format = number_format
            ws[f'C{row}'].border = border
            ws[f'C{row}'].number_format = currency_format
            ws[f'D{row}'].border = border
            ws[f'D{row}'].number_format = currency_format

            subtotal_semana += product_data['total']
            row += 1

        # Total de semana
        ws[f'A{row}'] = "TOTAL SEMANA"
        ws[f'A{row}'].font = total_font
        ws[f'A{row}'].fill = total_fill
        ws[f'D{row}'] = subtotal_semana
        ws[f'D{row}'].font = total_font
        ws[f'D{row}'].fill = total_fill
        ws[f'D{row}'].number_format = currency_format
        ws[f'D{row}'].border = border

        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = total_fill
            ws.cell(row=row, column=col).border = border

        row += 2

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Guardar en bytes
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_orders_by_client_to_excel(orders_by_client: dict, month: str) -> bytes:
    """
    Exportar pedidos agrupados por cliente a Excel.

    Args:
        orders_by_client: Dict con estructura:
            {
                'client_name': [
                    {'id': 1, 'created_at': '...', 'display_total': 50000, 'items': [...]},
                    ...
                ],
                ...
            }
        month: Mes en formato 'YYYY-MM'

    Returns:
        Bytes del archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    client_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    client_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format = '#,##0'

    # Encabezado
    ws['A1'] = f"PEDIDOS DEL MES {month}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    row = 3

    # Iterar por clientes
    total_general = 0
    for client_name in sorted(orders_by_client.keys()):
        orders = orders_by_client[client_name]

        # Encabezado de cliente
        ws[f'A{row}'] = client_name
        ws[f'A{row}'].font = client_font
        ws[f'A{row}'].fill = client_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # Encabezados de tabla
        headers = ['Fecha', 'Productos', 'Cantidad items', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        row += 1

        # Datos de pedidos
        subtotal_cliente = 0
        for order in orders:
            ws[f'A{row}'] = order['created_at'][:10]  # Solo fecha

            # Listar productos
            products_str = ", ".join([item['product_name'] for item in order.get('items', [])])
            ws[f'B{row}'] = products_str

            ws[f'C{row}'] = len(order.get('items', []))
            ws[f'D{row}'] = order['display_total']

            # Aplicar estilos
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws[f'C{row}'].border = border
            ws[f'D{row}'].border = border
            ws[f'D{row}'].number_format = currency_format

            subtotal_cliente += order['display_total']
            row += 1

        # Subtotal cliente
        ws[f'A{row}'] = "SUBTOTAL"
        ws[f'D{row}'] = subtotal_cliente
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].number_format = currency_format
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border

        total_general += subtotal_cliente
        row += 2

    # Total general
    ws[f'A{row}'] = "TOTAL GENERAL"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'D{row}'] = total_general
    ws[f'D{row}'].font = Font(bold=True, size=11)
    ws[f'D{row}'].number_format = currency_format
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.cell(row=row, column=col).border = border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Guardar en bytes
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
